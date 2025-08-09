import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# --- Configuration ---
# Define the full path for your Excel file
excel_save_path = r"C:\Users\ERP\Desktop\Jap\Reports_and_Was"
excel_filename = "Daily_Work_Log.xlsx"
full_excel_path = os.path.join(excel_save_path, excel_filename)

# --- Function to get user input ---
def get_user_input():
    print("\n--- Enter Work Log Details ---")
    
    # Get today's date automatically
    log_date = datetime.now().strftime("%Y-%m-%d") # YYYY-MM-DD format for Excel sorting
    print(f"Date (auto-filled): {log_date}")

    # You can pre-define time slots or let the user enter freely
    time_slots = [
        "8:00am - 10:00am",
        "10:00am - 12:00pm",
        "1:00pm - 2:00pm",
        "2:00pm - 4:00pm",
        "4:00pm - 5:00pm"
    ]
    
    # Let the user pick from predefined time slots or enter custom
    selected_time_slot = ""
    while True:
        print("\nAvailable Time Slots:")
        for i, slot in enumerate(time_slots):
            print(f"{i+1}. {slot}")
        print(f"{len(time_slots)+1}. Custom Time Slot")
        
        choice = input("Enter number for time slot or 'q' to quit: ").strip().lower()
        if choice == 'q':
            return None # User wants to quit
        
        try:
            choice_idx = int(choice) - 1
            if 0 <= choice_idx < len(time_slots):
                selected_time_slot = time_slots[choice_idx]
                break
            elif choice_idx == len(time_slots): # Custom option
                selected_time_slot = input("Enter custom time slot (e.g., '3:00pm - 3:30pm'): ").strip()
                if not selected_time_slot:
                    print("Custom time slot cannot be empty. Please try again.")
                    continue
                break
            else:
                print("Invalid choice. Please enter a valid number.")
        except ValueError:
            print("Invalid input. Please enter a number or 'q'.")

    work_description = input("Enter Work Done Description (type 'DONE' on a new line to finish multi-line input):\n")
    # For multi-line input:
    lines = []
    while True:
        line = input()
        if line.strip().upper() == 'DONE':
            break
        lines.append(line)
    work_description = "\n".join(lines)
    
    if not work_description.strip():
        print("Work description cannot be empty. Please re-enter.")
        return get_user_input() # Re-prompt if description is empty

    return {
        "Date": log_date,
        "Time Slot": selected_time_slot,
        "Work Done Description": work_description
    }

# --- Function to write data to Excel ---
def write_to_excel(data_row, file_path):
    # Ensure the directory exists
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    # Check if the file exists and load it, otherwise create a new workbook
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Add headers if it's a new file
        sheet.append(["Date", "Time Slot", "Work Done Description"])
        # Optional: Set column widths for new file
        sheet.column_dimensions['A'].width = 15 # Date
        sheet.column_dimensions['B'].width = 25 # Time Slot
        sheet.column_dimensions['C'].width = 60 # Work Done Description (adjust as needed)

    # Append the new row of data
    sheet.append([data_row["Date"], data_row["Time Slot"], data_row["Work Done Description"]])

    # Save the workbook
    try:
        workbook.save(file_path)
        print(f"Data successfully added to {file_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")

# --- Main execution ---
if __name__ == "__main__":
    while True:
        entry_data = get_user_input()
        if entry_data is None: # User chose to quit
            print("Exiting work log entry.")
            break
        
        write_to_excel(entry_data, full_excel_path)
        
        another_entry = input("\nDo you want to add another entry? (yes/no): ").strip().lower()
        if another_entry != 'yes':
            break

    print("Script finished.")